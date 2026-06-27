"""
Compare all three stance-detection methods against manual labels.
Run after: 04_keywords.py, 06_zeroshot.py, 07_finetune.py, and manual labeling done.
Run: python compare.py

Outputs:
  • data_processed/comparison_report.xlsx  — accuracy / F1 per method + confusion matrices
  • data_processed/plots/comparison/       — bar charts, confusion matrices, event analysis

Note on fairness:
  keyword & zero-shot are evaluated on ALL labeled rows.
  fine-tuned is evaluated on its out-of-fold (OOF) predictions — every labeled
  row was predicted by a CV fold that did NOT train on it, so all three methods
  are now compared on the same full labeled set with no leakage.
"""

import os
import pandas as pd
import numpy as np
import matplotlib
matplotlib.use("Agg")
import matplotlib.pyplot as plt
import seaborn as sns
from sklearn.metrics import (
    accuracy_score,
    f1_score,
    classification_report,
    confusion_matrix,
)
from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.utils.dataframe import dataframe_to_rows

PROC_DIR      = "data_processed"
VALID_CSV     = os.path.join(PROC_DIR, "validation_sample_1000.csv")
STANCE_OUT    = os.path.join(PROC_DIR, "messages_plus_sentiment_keywords.parquet")
ZEROSHOT_OUT  = os.path.join(PROC_DIR, "messages_clean_plus_zeroshot.parquet")
FINETUNE_OUT  = os.path.join(PROC_DIR, "messages_plus_all_plus_finetuned.parquet")
REPORT_XLSX   = os.path.join(PROC_DIR, "comparison_report.xlsx")
FINETUNE_OOF  = os.path.join(PROC_DIR, "finetuned_oof_predictions.csv")
PLOTS_DIR     = os.path.join(PROC_DIR, "plots", "comparison")
EVENTS_DIR    = os.path.join(PLOTS_DIR, "events")

# ── Key events ────────────────────────────────────────────────────────────────
# Each entry: (date_str YYYY-MM-DD, short label for plots, description for Excel)
EVENTS = [
    # Israel-Palestine conflict milestones
    ("2023-10-07", "Oct7_Hamas_Attack",        "Hamas attack on Israel / Nova festival massacre"),
    ("2023-10-17", "Oct17_Hospital",           "Al-Ahli hospital explosion, Gaza"),
    ("2023-11-24", "Nov24_Ceasefire_Deal",     "First hostage-ceasefire deal"),
    ("2024-05-07", "May7_Rafah_Operation",     "Israeli ground operation begins in Rafah"),
    ("2024-11-21", "Nov21_ICC_Warrant",        "ICC arrest warrant issued for Netanyahu"),
    # Israel-Iran direct escalation
    ("2024-04-01", "Apr1_Damascus_Strike",     "Israel strikes Iranian consulate in Damascus, kills IRGC generals"),
    ("2024-04-13", "Apr13_Iran_TruePromise1",  "Iran Operation True Promise: ~300 drones & missiles at Israel"),
    ("2024-04-19", "Apr19_Israel_Strike_Iran", "Israel retaliatory strike on Iran near Isfahan"),
    ("2024-10-01", "Oct1_Iran_TruePromise2",   "Iran Operation True Promise 2: ~180 ballistic missiles at Israel"),
    ("2024-10-25", "Oct25_Israel_Strike_Iran2","Israel strikes Iranian air defense & military sites"),
    # ── 2025 ──────────────────────────────────────────────────────────────────
    ("2025-06-13", "Jun13_TwelveDayWar_Start",  "Twelve Day War: Israel-Iran direct war begins (Jun 13–24, 2025)"),
    # ── 2026 ──────────────────────────────────────────────────────────────────
    ("2026-02-28", "Feb28_Israel_USA_Attack_Iran", "Israel & US launch war against Iran (Feb 28, 2026)"),
    ("2026-03-02", "Mar02_Iran_Response",           "Iran's response to Israel/US attack (Feb 28–Mar 5, 2026) — midpoint anchor"),
]

WINDOW_DAYS = 14   # days before and after each event to analyse

LABELS = ["pro_palestine", "pro_israel", "neutral"]

METHODS = {
    "keywords":  ("stance_label",    STANCE_OUT),
    "zeroshot":  ("stance_zeroshot", ZEROSHOT_OUT),
    "finetuned": ("stance_finetuned", FINETUNE_OUT),
}


def load_predictions(valid_df, method_col, parquet_path):
    """Join validation labels with predictions from the full-dataset parquet.
    Used for keyword & zero-shot, which never trained on the labeled rows."""
    if not os.path.exists(parquet_path):
        return None

    df = pd.read_parquet(parquet_path, columns=["row_id", method_col])
    merged = valid_df.merge(df, on="row_id", how="left").dropna(subset=[method_col])
    if len(merged) == 0:
        return None
    return merged["manual_label"].tolist(), merged[method_col].tolist()


def load_finetuned_oof(valid_df):
    """Out-of-fold predictions from finetune.py: every labeled row was predicted
    by a CV fold that did NOT train on it → leakage-free over ALL labeled rows."""
    if not os.path.exists(FINETUNE_OOF):
        return None
    oof = pd.read_csv(FINETUNE_OOF)  # columns: row_id, stance_finetuned_oof
    merged = valid_df.merge(oof, on="row_id", how="inner") \
                     .dropna(subset=["stance_finetuned_oof"])
    if len(merged) == 0:
        return None
    return merged["manual_label"].tolist(), merged["stance_finetuned_oof"].tolist()


def metrics_df(y_true, y_pred, method_name):
    acc = accuracy_score(y_true, y_pred)
    f1  = f1_score(y_true, y_pred, labels=LABELS, average="macro", zero_division=0)
    rows = []
    report = classification_report(
        y_true, y_pred, labels=LABELS, output_dict=True, zero_division=0
    )
    for label in LABELS:
        r = report.get(label, {})
        rows.append({
            "method":    method_name,
            "class":     label,
            "precision": round(r.get("precision", 0), 3),
            "recall":    round(r.get("recall", 0), 3),
            "f1":        round(r.get("f1-score", 0), 3),
            "support":   int(r.get("support", 0)),
        })
    summary = {
        "method":    method_name,
        "class":     "OVERALL",
        "precision": "",
        "recall":    "",
        "f1":        round(f1, 3),
        "support":   len(y_true),
    }
    rows.append(summary)
    return acc, f1, pd.DataFrame(rows)


def plot_overall(summary_rows):
    os.makedirs(PLOTS_DIR, exist_ok=True)
    df = pd.DataFrame(summary_rows)
    fig, axes = plt.subplots(1, 2, figsize=(10, 4))

    axes[0].bar(df["method"], df["accuracy"], color=["steelblue", "coral", "seagreen"])
    axes[0].set_ylim(0, 1)
    axes[0].set_title("Accuracy by method")
    axes[0].set_ylabel("Accuracy")
    for i, v in enumerate(df["accuracy"]):
        axes[0].text(i, v + 0.01, f"{v:.3f}", ha="center", fontsize=9)

    axes[1].bar(df["method"], df["f1_macro"], color=["steelblue", "coral", "seagreen"])
    axes[1].set_ylim(0, 1)
    axes[1].set_title("Macro F1 by method")
    axes[1].set_ylabel("F1 (macro)")
    for i, v in enumerate(df["f1_macro"]):
        axes[1].text(i, v + 0.01, f"{v:.3f}", ha="center", fontsize=9)

    plt.tight_layout()
    path = os.path.join(PLOTS_DIR, "overall_comparison.png")
    plt.savefig(path, dpi=150)
    plt.close()
    print(f"Saved plot → {path}")


def plot_confusion(y_true, y_pred, method_name):
    os.makedirs(PLOTS_DIR, exist_ok=True)
    cm = confusion_matrix(y_true, y_pred, labels=LABELS)
    fig, ax = plt.subplots(figsize=(6, 5))
    sns.heatmap(
        cm, annot=True, fmt="d", cmap="Blues",
        xticklabels=LABELS, yticklabels=LABELS, ax=ax,
    )
    ax.set_xlabel("Predicted")
    ax.set_ylabel("Actual")
    ax.set_title(f"Confusion matrix – {method_name}")
    plt.tight_layout()
    path = os.path.join(PLOTS_DIR, f"confusion_{method_name}.png")
    plt.savefig(path, dpi=150)
    plt.close()
    print(f"Saved plot → {path}")


def write_xlsx(summary_rows, detail_frames, event_df=None):
    wb = Workbook()
    ws_sum = wb.active
    ws_sum.title = "Summary"

    ws_sum["A1"] = "Method comparison – overall metrics"
    ws_sum["A1"].font = Font(bold=True, size=13)

    sum_df = pd.DataFrame(summary_rows)
    for r_idx, row in enumerate(dataframe_to_rows(sum_df, index=False, header=True), start=3):
        for c_idx, val in enumerate(row, start=1):
            ws_sum.cell(row=r_idx, column=c_idx, value=val)

    for method_name, df_detail in detail_frames.items():
        ws = wb.create_sheet(title=method_name)
        ws["A1"] = f"Per-class metrics – {method_name}"
        ws["A1"].font = Font(bold=True)
        for r_idx, row in enumerate(dataframe_to_rows(df_detail, index=False, header=True), start=3):
            for c_idx, val in enumerate(row, start=1):
                ws.cell(row=r_idx, column=c_idx, value=val)

    if event_df is not None and len(event_df) > 0:
        ws_ev = wb.create_sheet(title="Event_Analysis")
        ws_ev["A1"] = "Sentiment & stance before vs after key events"
        ws_ev["A1"].font = Font(bold=True, size=13)
        for r_idx, row in enumerate(dataframe_to_rows(event_df, index=False, header=True), start=3):
            for c_idx, val in enumerate(row, start=1):
                ws_ev.cell(row=r_idx, column=c_idx, value=val)

        ws_ev2 = wb.create_sheet(title="Events_List")
        ws_ev2["A1"] = "Key events analysed"
        ws_ev2["A1"].font = Font(bold=True)
        events_list_df = pd.DataFrame(EVENTS, columns=["date", "label", "description"])
        for r_idx, row in enumerate(dataframe_to_rows(events_list_df, index=False, header=True), start=3):
            for c_idx, val in enumerate(row, start=1):
                ws_ev2.cell(row=r_idx, column=c_idx, value=val)

    wb.save(REPORT_XLSX)
    print(f"Saved report → {REPORT_XLSX}")


def run_event_analysis():
    """
    For each event in EVENTS, analyse ±WINDOW_DAYS days of data.
    Produces per-event plots and a summary Excel sheet.
    """
    os.makedirs(EVENTS_DIR, exist_ok=True)

    # Use the best available stance method
    if os.path.exists(FINETUNE_OUT):
        source_file  = FINETUNE_OUT
        stance_col   = "stance_finetuned"
        method_used  = "fine-tuned (BERTweet)"
    elif os.path.exists(ZEROSHOT_OUT):
        source_file  = ZEROSHOT_OUT
        stance_col   = "stance_zeroshot"
        method_used  = "zero-shot (DeBERTa)"
    elif os.path.exists(STANCE_OUT):
        source_file  = STANCE_OUT
        stance_col   = "stance_label"
        method_used  = "keyword"
    else:
        print("[SKIP] Event analysis: no stance parquet found.")
        return None

    print(f"  Using stance method: {method_used}")
    df = pd.read_parquet(source_file)
    df["created_at"] = pd.to_datetime(df["created_at"], errors="coerce", utc=True)
    df = df.dropna(subset=["created_at"])
    df["date"] = df["created_at"].dt.date

    if len(df) == 0:
        print("[SKIP] Event analysis: no parseable timestamps.")
        return None

    summary_rows = []

    for date_str, label, description in EVENTS:
        event_date = pd.Timestamp(date_str, tz="UTC")
        start      = (event_date - pd.Timedelta(days=WINDOW_DAYS)).date()
        end        = (event_date + pd.Timedelta(days=WINDOW_DAYS)).date()

        window = df[(df["date"] >= start) & (df["date"] <= end)].copy()

        if len(window) < 10:
            print(f"  [SKIP] {label}: only {len(window)} rows in window — not enough data.")
            continue

        print(f"  {label}: {len(window):,} rows ({start} → {end})")

        window["day_offset"] = (
            pd.to_datetime(window["date"]) - pd.Timestamp(date_str)
        ).dt.days

        # ── daily stats by stance group ───────────────────────────────────────
        if "stance_label" in window.columns and stance_col != "stance_label":
            window = window.drop(columns=["stance_label"])
        window = window.rename(columns={stance_col: "stance_label"})

        daily = (
            window.groupby(["day_offset", "stance_label"])
            .agg(
                count=("text", "count"),
                pct_negative=("sentiment_label", lambda x: (x == "negative").mean() * 100),
            )
            .reset_index()
        )

        fig, axes = plt.subplots(2, 1, figsize=(12, 8), sharex=True)
        fig.suptitle(f"{label}\n{description}", fontsize=11, y=1.01)

        colors = {"pro_palestine": "green", "pro_israel": "royalblue", "neutral": "gray"}

        # Top: volume
        for stance, grp in daily.groupby("stance_label"):
            axes[0].plot(
                grp["day_offset"], grp["count"],
                label=stance, color=colors.get(stance, "black"), marker="o", markersize=3,
            )
        axes[0].axvline(0, color="red", linestyle="--", linewidth=1.2, label="Event date")
        axes[0].set_ylabel("Messages per day")
        axes[0].set_title("Daily message volume by stance")
        axes[0].legend(fontsize=8)

        # Bottom: % negative sentiment
        for stance, grp in daily.groupby("stance_label"):
            axes[1].plot(
                grp["day_offset"], grp["pct_negative"],
                label=stance, color=colors.get(stance, "black"), marker="o", markersize=3,
            )
        axes[1].axvline(0, color="red", linestyle="--", linewidth=1.2, label="Event date")
        axes[1].set_ylabel("% negative sentiment")
        axes[1].set_xlabel(f"Days relative to event  (0 = {date_str})")
        axes[1].set_title("Daily % negative sentiment by stance")
        axes[1].legend(fontsize=8)

        plt.tight_layout()
        plot_path = os.path.join(EVENTS_DIR, f"{label}.png")
        plt.savefig(plot_path, dpi=150, bbox_inches="tight")
        plt.close()
        print(f"    Saved → {plot_path}")

        # ── summary stats: before vs after ───────────────────────────────────
        before = window[window["day_offset"] < 0]
        after  = window[window["day_offset"] > 0]

        for period_label, period_df in [("before", before), ("after", after)]:
            if len(period_df) == 0:
                continue
            for stance in ["pro_palestine", "pro_israel", "neutral"]:
                sub = period_df[period_df["stance_label"] == stance]
                summary_rows.append({
                    "event":         label,
                    "date":          date_str,
                    "description":   description,
                    "stance_method": method_used,
                    "period":        period_label,
                    "stance":        stance,
                    "n_messages":    len(sub),
                    "pct_negative":  round((sub["sentiment_label"] == "negative").mean() * 100, 1) if len(sub) else None,
                    "pct_positive":  round((sub["sentiment_label"] == "positive").mean() * 100, 1) if len(sub) else None,
                    "pct_neutral":   round((sub["sentiment_label"] == "neutral").mean() * 100, 1)  if len(sub) else None,
                })

    if not summary_rows:
        print("  No events had sufficient data.")
        return None

    summary_df = pd.DataFrame(summary_rows)
    print(f"\nEvent analysis complete — {len(EVENTS)} events processed.")
    return summary_df


def main():
    # ── 1. Load manual labels ───────────────────────────────────────────────
    if not os.path.exists(VALID_CSV):
        raise SystemExit(f"Missing: {VALID_CSV}\nRun 05_make_validation_sample.py first.")

    valid_df = pd.read_csv(VALID_CSV)
    valid_df["manual_label"] = valid_df["manual_label"].astype(str).str.strip()
    valid_df = valid_df[valid_df["manual_label"].isin(LABELS)].copy()

    if len(valid_df) < 50:
        raise SystemExit(
            f"Only {len(valid_df)} labeled rows. Fill in more manual_label values first."
        )

    print(f"Labeled rows: {len(valid_df)}")
    print(valid_df["manual_label"].value_counts().to_string())

    # ── 2. Compute metrics per method ───────────────────────────────────────
    # keyword & zero-shot: evaluated on all labeled rows (never trained on them).
    # fine-tuned: evaluated on out-of-fold predictions (leakage-free, all rows).
    summary_rows  = []
    detail_frames = {}
    all_preds     = {}

    for method_name, (col, parquet_path) in METHODS.items():
        if method_name == "finetuned":
            result = load_finetuned_oof(valid_df)
            if result is None:
                print(f"\n[SKIP] {method_name}: {FINETUNE_OOF} not found. Run finetune.py.")
                continue
        else:
            result = load_predictions(valid_df, col, parquet_path)
            if result is None:
                print(f"\n[SKIP] {method_name}: {parquet_path} not found.")
                continue

        y_true, y_pred = result
        if len(y_true) == 0:
            print(f"\n[SKIP] {method_name}: no matching rows after merge.")
            continue

        acc, f1, df_detail = metrics_df(y_true, y_pred, method_name)
        summary_rows.append({"method": method_name, "accuracy": round(acc, 3), "f1_macro": round(f1, 3)})
        detail_frames[method_name] = df_detail
        all_preds[method_name]     = (y_true, y_pred)

        print(f"\n── {method_name} ──")
        print(f"  Evaluated on : {len(y_true)} labeled rows")
        print(f"  Accuracy : {acc:.3f}")
        print(f"  F1 macro : {f1:.3f}")
        print(classification_report(y_true, y_pred, labels=LABELS, zero_division=0))

    if not summary_rows:
        raise SystemExit("No method produced results. Make sure the parquet files exist.")

    # ── 3. Plots ────────────────────────────────────────────────────────────
    plot_overall(summary_rows)
    for method_name, (y_true, y_pred) in all_preds.items():
        plot_confusion(y_true, y_pred, method_name)

    # ── 4. Event analysis ───────────────────────────────────────────────────
    print("\n── Event analysis ──")
    event_df = run_event_analysis()

    # ── 5. Excel report ─────────────────────────────────────────────────────
    write_xlsx(summary_rows, detail_frames, event_df)


if __name__ == "__main__":
    main()